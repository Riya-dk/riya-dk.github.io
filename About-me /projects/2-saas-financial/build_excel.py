"""Build Excel dashboard with pivot-style summaries and charts."""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent
DATA = ROOT.parent.parent / "project-datasources" / "2-saas-financial" / "data"
OUT = ROOT / "SaaS_Financial_Dashboard.xlsx"

def main():
    sub = pd.read_csv(DATA / "FactSubscriptions.csv")
    plan = pd.read_csv(DATA / "DimPlan.csv")
    costs = pd.read_csv(DATA / "FactCosts.csv")
    sub = sub.merge(plan[["PlanID", "PlanName"]], on="PlanID")

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Subscriptions"
    ws_data.append(["CustomerID", "PlanID", "Segment", "PlanName", "MRR"])
    for _, r in sub.iterrows():
        ws_data.append([r["CustomerID"], r["PlanID"], r["Segment"], r["PlanName"], r["MRR"]])

    by_plan = sub.groupby("PlanName").agg(MRR=("MRR", "sum")).reset_index()
    by_seg = sub.groupby("Segment").agg(MRR=("MRR", "sum")).reset_index()

    ws_plan = wb.create_sheet("MRR by Plan")
    ws_plan.append(["Plan", "MRR"])
    for _, r in by_plan.iterrows():
        ws_plan.append([r["PlanName"], r["MRR"]])

    ws_seg = wb.create_sheet("MRR by Segment")
    ws_seg.append(["Segment", "MRR"])
    for _, r in by_seg.iterrows():
        ws_seg.append([r["Segment"], r["MRR"]])

    wb.save(OUT)
    print("Saved", OUT)

if __name__ == "__main__":
    main()
